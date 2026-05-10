"""
Lead Generation Automation Script
====================================
Collects nonprofit lead data from a public API + curated seed dataset,
enriches records with generated emails & LinkedIn URLs, cleans the data,
and exports a formatted multi-sheet Excel workbook.

Tools  : requests, pandas, openpyxl, schedule
Run    : python lead_scraper.py
Schedule: python lead_scraper.py --schedule 24   (every 24 h)
"""

import re
import sys
import time
import logging
import requests
import schedule
import pandas as pd
from pathlib import Path
from datetime import datetime

# ── Setup ─────────────────────────────────────────────────────────────────────

Path("output").mkdir(exist_ok=True)
Path("logs").mkdir(exist_ok=True)

logging.basicConfig(
    filename="logs/run.log",
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)
OUTPUT_DIR = Path("output")


# ── 1. Data Collection ────────────────────────────────────────────────────────

def fetch_leads() -> list:
    """
    Primary  : Open Data Soft US-Nonprofits public dataset (no API key).
    Fallback : Curated seed list of 30 well-known nonprofits.
    Always ensures >= 30 final records.
    """
    leads = []

    url = (
        "https://data.opendatasoft.com/api/explore/v2.1/catalog"
        "/datasets/us-nonprofits/records"
    )
    params = {
        "limit": 30,
        "select": "organization_name,city,state,ntee_code,asset_amount,website",
        "where": "city IS NOT NULL AND organization_name IS NOT NULL",
        "order_by": "asset_amount DESC",
    }
    try:
        logger.info("Requesting live API data...")
        r = requests.get(url, params=params, timeout=15)
        r.raise_for_status()
        for rec in r.json().get("results", []):
            leads.append({
                "Name":       rec.get("organization_name", "").strip().title(),
                "City":       rec.get("city", "").strip().title(),
                "State":      rec.get("state", "").strip().upper(),
                "Location":   "{}, {}".format(
                    rec.get("city", "").title(),
                    rec.get("state", "").upper()
                ),
                "Website":    rec.get("website") or "",
                "Sector":     _ntee_sector(rec.get("ntee_code", "")),
                "Assets_USD": rec.get("asset_amount") or 0,
                "Email":      "",
                "LinkedIn":   "",
                "Source":     "API",
            })
        logger.info("API returned {} records.".format(len(leads)))
    except Exception as exc:
        logger.warning("API unavailable ({}). Using seed data.".format(exc))

    # Pad to >= 30 with seed records
    if len(leads) < 30:
        seen = {r["Name"] for r in leads}
        for row in _seed():
            if row["Name"] not in seen:
                leads.append(row)
                seen.add(row["Name"])
            if len(leads) >= 30:
                break

    return leads


def _ntee_sector(code):
    """Converts an NTEE letter code to a readable sector name."""
    tbl = {
        "A": "Arts & Culture",    "B": "Education",         "C": "Environment",
        "D": "Animal Welfare",    "E": "Health",             "F": "Mental Health",
        "G": "Disease Research",  "H": "Medical Research",   "I": "Crime & Safety",
        "J": "Employment",        "K": "Food & Agriculture", "L": "Housing",
        "M": "Public Safety",     "N": "Recreation",         "O": "Youth Development",
        "P": "Human Services",    "Q": "International",      "R": "Civil Rights",
        "S": "Community Dev.",    "T": "Philanthropy",        "U": "Science",
        "V": "Social Science",    "W": "Public Affairs",     "X": "Religion",
    }
    return tbl.get((code or "?")[0].upper(), "Other")


def _seed():
    """30 well-known nonprofits used as fallback / supplement."""
    rows = [
        ("Doctors Without Borders USA",      "New York",      "NY", "https://www.doctorswithoutborders.org", "Health",              520_000_000),
        ("American Red Cross",               "Washington",    "DC", "https://www.redcross.org",               "Human Services",    3_100_000_000),
        ("Feeding America",                  "Chicago",       "IL", "https://www.feedingamerica.org",          "Food & Agriculture",   280_000_000),
        ("Habitat for Humanity",             "Atlanta",       "GA", "https://www.habitat.org",                "Housing",              430_000_000),
        ("World Wildlife Fund",              "Washington",    "DC", "https://www.worldwildlife.org",           "Environment",          380_000_000),
        ("Sierra Club Foundation",           "San Francisco", "CA", "https://www.sierraclubfoundation.org",   "Environment",          120_000_000),
        ("ACLU Foundation",                  "New York",      "NY", "https://www.aclu.org",                   "Civil Rights",         340_000_000),
        ("Boys and Girls Clubs of America",  "Atlanta",       "GA", "https://www.bgca.org",                   "Youth Development",    210_000_000),
        ("St. Jude Childrens Hospital",      "Memphis",       "TN", "https://www.stjude.org",                 "Medical Research",   7_200_000_000),
        ("United Way Worldwide",             "Alexandria",    "VA", "https://www.unitedway.org",              "Human Services",       190_000_000),
        ("Save the Children Federation",     "Fairfield",     "CT", "https://www.savethechildren.org",        "International",        540_000_000),
        ("The Nature Conservancy",           "Arlington",     "VA", "https://www.nature.org",                 "Environment",        6_800_000_000),
        ("Goodwill Industries International","Rockville",     "MD", "https://www.goodwill.org",               "Employment",           310_000_000),
        ("Planned Parenthood Federation",    "New York",      "NY", "https://www.plannedparenthood.org",      "Health",               480_000_000),
        ("Smithsonian Institution",          "Washington",    "DC", "https://www.si.edu",                     "Arts & Culture",     2_900_000_000),
        ("Alzheimers Association",           "Chicago",       "IL", "https://www.alz.org",                    "Disease Research",     320_000_000),
        ("National Public Radio",            "Washington",    "DC", "https://www.npr.org",                    "Arts & Culture",       640_000_000),
        ("American Cancer Society",          "Atlanta",       "GA", "https://www.cancer.org",                 "Disease Research",   1_700_000_000),
        ("CARE USA",                         "Atlanta",       "GA", "https://www.care.org",                   "International",        110_000_000),
        ("National Geographic Society",      "Washington",    "DC", "https://www.nationalgeographic.org",     "Science",            1_100_000_000),
        ("Food Bank for New York City",      "New York",      "NY", "https://www.foodbanknyc.org",            "Food & Agriculture",    95_000_000),
        ("American Heart Association",       "Dallas",        "TX", "https://www.heart.org",                  "Health",             2_200_000_000),
        ("Oxfam America",                    "Boston",        "MA", "https://www.oxfamamerica.org",            "International",         88_000_000),
        ("Khan Academy",                     "Mountain View", "CA", "https://www.khanacademy.org",             "Education",            340_000_000),
        ("Wikimedia Foundation",             "San Francisco", "CA", "https://wikimediafoundation.org",         "Education",            260_000_000),
        ("Direct Relief",                    "Santa Barbara", "CA", "https://www.directrelief.org",            "Health",               870_000_000),
        ("Teach For America",                "New York",      "NY", "https://www.teachforamerica.org",         "Education",            490_000_000),
        ("Environmental Defense Fund",       "New York",      "NY", "https://www.edf.org",                    "Environment",          960_000_000),
        ("Prison Policy Initiative",         "Northampton",   "MA", "https://www.prisonpolicy.org",            "Civil Rights",           5_000_000),
        ("Comic Relief US",                  "New York",      "NY", "https://www.comicrelief.com/usa",         "Philanthropy",          45_000_000),
    ]
    return [
        {
            "Name": n, "City": c, "State": s,
            "Location": "{}, {}".format(c, s),
            "Website": w, "Sector": sec, "Assets_USD": a,
            "Email": "", "LinkedIn": "", "Source": "Seed",
        }
        for n, c, s, w, sec, a in rows
    ]


# ── 2. Email & LinkedIn Generation (Bonus Feature) ────────────────────────────

_VALID_EMAIL = re.compile(r"^[\w.\-]+@[\w.\-]+\.\w{2,}$")


def generate_email(website):
    """Derives info@<domain> from an organisation's website URL."""
    if not website:
        return ""
    domain = re.sub(r"https?://", "", website).split("/")[0].lower()
    domain = re.sub(r"^www\.", "", domain)
    return "info@{}".format(domain) if ("." in domain and domain) else ""


def generate_linkedin(name):
    """Creates a probable LinkedIn /company/ slug from the org name."""
    slug = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    return "https://www.linkedin.com/company/{}".format(slug)


# ── 3. Data Cleaning ──────────────────────────────────────────────────────────

def clean_data(df):
    """
    Cleaning steps:
      1. Strip whitespace from all string columns
      2. Remove duplicate rows (keyed on Name + Location)
      3. Replace empty display fields with 'N/A'
      4. Coerce Assets_USD to int, fill missing with 0
      5. Add Email_Valid flag (Yes / No)
      6. Sort by Sector then Name
    """
    logger.info("Cleaning {} raw rows...".format(len(df)))

    for col in df.select_dtypes(include=["object"]).columns:
        df[col] = df[col].str.strip()

    before = len(df)
    df.drop_duplicates(subset=["Name", "Location"], keep="first", inplace=True)
    logger.info("Dropped {} duplicates.".format(before - len(df)))

    for col in ["Website", "Email", "LinkedIn"]:
        df[col] = df[col].replace("", "N/A")
    df["Sector"] = df["Sector"].fillna("Unknown")
    df["Assets_USD"] = (
        pd.to_numeric(df["Assets_USD"], errors="coerce").fillna(0).astype(int)
    )

    df["Email_Valid"] = df["Email"].apply(
        lambda e: "Yes" if e != "N/A" and _VALID_EMAIL.match(e) else "No"
    )

    df.sort_values(["Sector", "Name"], inplace=True)
    df.reset_index(drop=True, inplace=True)
    logger.info("Clean dataset: {} rows.".format(len(df)))
    return df


# ── 4. Excel Export ───────────────────────────────────────────────────────────

def export_excel(df, filename):
    """
    Exports to a styled 2-sheet .xlsx workbook:
      'Leads'   — full dataset, alternating row colours, auto-width columns
      'Summary' — key metrics (total leads, email coverage, sector count, etc.)
    """
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    path = OUTPUT_DIR / filename

    with pd.ExcelWriter(path, engine="openpyxl") as writer:

        # ── Leads sheet ──────────────────────────────────────────────────────
        df.to_excel(writer, index=False, sheet_name="Leads")
        ws = writer.sheets["Leads"]

        h_fill   = PatternFill("solid", fgColor="2E75B6")
        alt_fill = PatternFill("solid", fgColor="EBF3FB")
        na_fill  = PatternFill("solid", fgColor="FFDADA")
        h_font   = Font(bold=True, color="FFFFFF", size=11)
        side     = Side(style="thin", color="CCCCCC")
        bdr      = Border(left=side, right=side, top=side, bottom=side)

        for ci in range(1, len(df.columns) + 1):
            c = ws.cell(row=1, column=ci)
            c.fill      = h_fill
            c.font      = h_font
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        for ri in range(2, len(df) + 2):
            for ci in range(1, len(df.columns) + 1):
                c = ws.cell(row=ri, column=ci)
                c.border    = bdr
                c.alignment = Alignment(vertical="center")
                val = str(c.value or "")
                if val in ("N/A", "0", ""):
                    c.fill = na_fill
                elif ri % 2 == 0:
                    c.fill = alt_fill

        for ci, col_name in enumerate(df.columns, 1):
            max_w = max(
                len(str(col_name)),
                *(len(str(ws.cell(row=r, column=ci).value or ""))
                  for r in range(2, len(df) + 2)),
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_w + 4, 50)

        ws.freeze_panes = "A2"

        # ── Summary sheet ────────────────────────────────────────────────────
        summary = pd.DataFrame({
            "Metric": [
                "Total Leads",
                "Leads with Generated Email",
                "Leads with Website",
                "Leads with LinkedIn URL",
                "Unique Sectors",
                "Data Sources",
                "Generated At",
            ],
            "Value": [
                len(df),
                int((df["Email"] != "N/A").sum()),
                int((df["Website"] != "N/A").sum()),
                int((df["LinkedIn"] != "N/A").sum()),
                df["Sector"].nunique(),
                ", ".join(df["Source"].unique()),
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ],
        })
        summary.to_excel(writer, index=False, sheet_name="Summary")
        ws2 = writer.sheets["Summary"]
        for c in ws2[1]:
            c.fill = h_fill
            c.font = h_font
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 38

    logger.info("Exported -> {}".format(path))
    return path


# ── 5. Main Pipeline ──────────────────────────────────────────────────────────

def run_pipeline():
    """Full end-to-end lead generation pipeline."""
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    divider = "=" * 54
    print("\n{}".format(divider))
    print("  Lead Generation Automation  |  {}".format(now))
    print(divider)

    print("\n[1/4] Collecting lead data...")
    raw = fetch_leads()
    df  = pd.DataFrame(raw)
    print("      OK  {} raw records fetched.".format(len(df)))

    print("\n[2/4] Enriching: generating emails & LinkedIn URLs...")
    df["Email"]    = df["Website"].apply(generate_email)
    df["LinkedIn"] = df["Name"].apply(generate_linkedin)
    enriched = (df["Email"] != "").sum()
    print("      OK  Emails generated for {}/{} leads.".format(enriched, len(df)))

    print("\n[3/4] Cleaning data...")
    df = clean_data(df)
    print("      OK  {} clean records ready.".format(len(df)))

    stamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = "leads_{}.xlsx".format(stamp)
    print("\n[4/4] Exporting -> output/{}".format(filename))
    out = export_excel(df, filename)
    print("      OK  Saved successfully.\n")

    print("-- Sample (first 5 rows) " + "-" * 29)
    print(df[["Name", "Location", "Sector", "Email"]].head(5).to_string(index=False))
    print("\nDone.  Output: {}\n".format(out))
    return out


# ── 6. Scheduler (Bonus Feature) ─────────────────────────────────────────────

def start_scheduler(hours=24):
    """Runs the full pipeline now, then repeats every `hours` hours."""
    print("Scheduler started -- pipeline will re-run every {} hour(s).".format(hours))
    print("Press Ctrl+C to stop.\n")
    schedule.every(hours).hours.do(run_pipeline)
    run_pipeline()
    while True:
        schedule.run_pending()
        time.sleep(60)


# ── Entry Point ───────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if "--schedule" in sys.argv:
        try:
            h = int(sys.argv[sys.argv.index("--schedule") + 1])
        except (IndexError, ValueError):
            h = 24
        start_scheduler(h)
    else:
        run_pipeline()
